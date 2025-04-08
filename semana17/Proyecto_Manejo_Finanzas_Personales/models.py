import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

excel_file = 'PFS.xlsx'
sheet_name_expense_income = 'Expense and incomes'
sheet_name_category = 'category'

def check_input_is_valid_string(parameter):

    if len(parameter) == 0 :
        res = 'Empty'
    else:
        res = 'Not empty'
    return res

def check_if_user_entry_has_spaces(parameter):
    
    try:
        parameter_value = str(parameter)
        print("User input is string")
        if parameter_value.isspace():
            res = 'has spaces'
        else:
            res = 'No spaces'
    except ValueError:
        print("User input is not an string")
        res = 'No string value'
    
    return res

class Category():

    def __init__(self, nombre, tipo):
        self.nombre = nombre
        self.tipo = tipo
    
    def save_category(parameter):

        try:
            workbook = load_workbook(excel_file)
        except FileNotFoundError:
            print('Error')
        if sheet_name_category in workbook.sheetnames:
            sheet = workbook[sheet_name_category]
        else:
            sheet = workbook.create_sheet(sheet_name_category)
        if sheet["A1"].value is None:    
            sheet["A1"] = 'Category'
            sheet["B1"] = 'Type'    
        sheet.append([parameter.nombre, parameter.tipo])  
        workbook.save(excel_file)
        workbook.close()

    def check_if_category_already_exist(category_parameter,list_category):

        for item in list_category:
            print(f'category parameter: {category_parameter[0]}, item of the list: {''.join(item[0])}')
            if category_parameter[0] == ''.join(item[0]):
                res = 'exist'
                break
            else:    
                res = 'does not exist'
        
        return res

class Transaction():

    def __init__(self, detail_p,category_p,type_p,amount_p):
        self.detail = detail_p
        self.category = category_p
        self.type = type_p
        self.amount = amount_p

    def save_transaction(parameter):

        try:
            workbook = load_workbook(excel_file)
        except FileNotFoundError:
            print('Error')   
        
        if sheet_name_expense_income in workbook.sheetnames:
            sheet = workbook[sheet_name_expense_income]
        else:
            sheet = workbook.create_sheet(sheet_name_expense_income)

        if sheet["A1"].value is None:
            
            sheet["A1"] = 'Detail'
            sheet["B1"] = 'Category'
            sheet["C1"] = 'Type'
            sheet["D1"] = 'Amount'   

        sheet.append([parameter.detail,parameter.category,parameter.type,parameter.amount])  
        
        workbook.save(excel_file)
        workbook.close()
    
    def check_valid_input_number(parameter):

        try:
            parameter_value = int(parameter)
            print("User input is an integer")
            if parameter_value >= 0:
                res = 'is numeric'
            else:
                res = 'Not numeric'
        except ValueError:
            print("User input is not an integer")
            res = 'Not numeric'
        
        return res
