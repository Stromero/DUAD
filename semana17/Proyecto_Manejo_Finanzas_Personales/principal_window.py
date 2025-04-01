import FreeSimpleGUI as sg
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import xlsxwriter

#Declare the elements
excel_file = 'PFS.xlsx'
sheet_name_expense_income = 'Expense and incomes'
sheet_name_category = 'category'
table_header_main_window = ['Detail','Category','Type','Amount']
table_header_category = ['Category']

def load_category_table():

    try:
        data = pd.read_excel(excel_file,sheet_name_category)
        return data.values.tolist()
    except Exception as e:
        print(f"Error when try to upload file: {e}")
        return None

def check_if_category_exist(category_parameter,list_category):

    for item in list_category:
        print(''.join(item))
        if category_parameter == ''.join(item):
            res = 'exist'
            break
        else:    
            res = 'does not exist'
    
    return res


def save_category(category_parameter):

    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        print('Error')
    if sheet_name_category in workbook.sheetnames:
        sheet = workbook[sheet_name_category]
    else:
        sheet = workbook.create_sheet(sheet_name_category)

    if sheet["A1"].value is None:    
        sheet["A1"] = 'Category'
        
    sheet.append([category_parameter])  
    
    workbook.save(excel_file)
    workbook.close()
    

def register(detail_parameter,category_parameter, type_parameter, amount_parameter):

    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        print('Error')   
    
    if sheet_name_expense_income in workbook.sheetnames:
        sheet = workbook[sheet_name_expense_income]
    else:
        sheet = workbook.create_sheet(sheet_name_expense_income)

    if sheet["A1"].value is None:
        
        sheet["A1"] = 'Detail'
        sheet["B1"] = 'Category'
        sheet["C1"] = 'Type'
        sheet["D1"] = 'Amount'   

    sheet.append([detail_parameter,category_parameter,type_parameter,amount_parameter])  
    
    workbook.save(excel_file)
    workbook.close()

def load_table():

    if os.path.exists(excel_file):
        print('existe')
        try:
                data = pd.read_excel(excel_file,sheet_name_expense_income)
                return data.values.tolist()
        except Exception as e:
                print(f"Error when try to upload file: {e}")
        return None
    else:
        print('No existe')
        wb = Workbook()
        wb.save(excel_file)


def main_page():
    
    layout = [
        [sg.Text('Personal financial management')],
        [sg.Text('My movements')],
        [sg.Table(values=table_values,headings=table_header_main_window,key='-MAINTABLE-')],
        [sg.Button('Add category', key='-ADDCATEGORY-'), sg.Button('Register expense', key='-EXPENSE-'), sg.Button('Register income', key='-INCOME-')],
        [sg.Button('Exit')]
    ]

    return sg.Window('PFM System',layout)

def add_category():

    layout = [
        [sg.Text('Add the new category:')],
        [sg.InputText(key='-INCATEGORY-')],
        [sg.Table(values=table_values_category, headings=table_header_category ,key='-CATEGORYTABLE-',
                auto_size_columns=False,justification='left',enable_events=True,num_rows=10)],
        [sg.Button('Add',key='-ADDCATEGORY-'),sg.Button('Cancel',key='-CANCELCATEGORY-'),sg.Button('Back')]
    ]

    return sg.Window('Category',layout)

def add_expense():
    
    layout = [
        [sg.Text('Add the new expense')],
        [sg.Text('Category: '), sg.Combo(values=table_values_category,key='-COMBOBOXCATEGORY-')],
        [sg.Text('Detail: '),sg.InputText(key='-INDETAIL-')],
        [sg.Text('Amount: '),sg.InputText(key='-AMOUNTDETAIL-')],
        [sg.Button('Add',key='-ADDEXPENSE-'),sg.Button('Cancel', key='-CANCELEXPENSE-'),sg.Button('Back')]
    ]

    return sg.Window('Expense',layout)
    

def add_income():
    
    layout = [
        [sg.Text('Add the income')],
        [sg.Text('Category: '),sg.Combo(values=table_values_category,key='-COMBOCATEGORY-')],
        [sg.Text('Detail: '),sg.InputText(key='-INDETAIL-')],
        [sg.Text('Amount: '),sg.InputText(key='-INAMOUNT-')],
        [sg.Button('Add',key='-ADDINCOME-'),sg.Button('Cancel',key='-CANCELINCOME-'),sg.Button('Back')]
    ]

    return sg.Window('Income',layout)

table_data = load_table()
if table_data is None:
    table_values = list()   
else:
    table_values = table_data.copy()

table_data = load_category_table()
if table_data is None:
    table_values_category = list()    
else:
    table_values_category = table_data.copy()

# Create the window
# Event loop to process and get the values from inputs

def main():

    window = main_page()

    while True:
        event, values = window.read()

        if event in (sg.WIN_CLOSED,'Exit'):
            break
        if event == '-ADDCATEGORY-':
            window.hide()
            new_window = add_category()
        if event == '-EXPENSE-':
            window.hide()
            new_window = add_expense()
        if event == '-INCOME-':
            window.hide()
            new_window = add_income()
        #else:
            #continue
        
        while True:
            new_event, new_values = new_window.read()

            if new_event in (sg.WIN_CLOSED, 'Back'):
                new_window.close()
                window.un_hide()
                break
            elif new_event == '-CANCELCATEGORY-':
                new_window['-INCATEGORY-'].update('')
            elif new_event == '-CANCELEXPENSE-':
                new_window['-COMBOBOXCATEGORY-'].update('')
                new_window['-INDETAIL-'].update('')
                new_window['-AMOUNTDETAIL-'].update('')
            elif new_event == '-CANCELINCOME-':
                new_window['-COMBOCATEGORY-'].update('')
                new_window['-INDETAIL-'].update('')
                new_window['-INAMOUNT-'].update('')
            elif new_event == '-ADDCATEGORY-':
                category_information = new_values['-INCATEGORY-'].strip()
                #print(f'value of category is: {category_information}')
                table_values_category.append([category_information])
                save_category(category_information)
                new_window['-INCATEGORY-'].update('')
                new_window['-CATEGORYTABLE-'].update(values=table_values_category)
            elif new_event == '-ADDEXPENSE-':
                category_information = new_values['-COMBOBOXCATEGORY-']
                convert_to_string = ' '.join(category_information)
                new_category_value = convert_to_string        
                res = check_if_category_exist(new_category_value,table_values_category)   
                #print(res)
                if res == 'does not exist':
                    sg.popup('Category does not exist, add the category to continue')
                    new_window.close()
                    window.un_hide()
                    break
                #print(f'new category info: {new_category_value}')
                detail_information = new_values['-INDETAIL-']
                amount_information = new_values['-AMOUNTDETAIL-']
                #print(f'the category selected is: {new_category_value}, detail information is: {detail_information}, amount of this expense is: {amount_information}')
                register(detail_information,new_category_value,'Expense',amount_information)
                new_window['-INDETAIL-'].update('')
                new_window['-AMOUNTDETAIL-'].update('')
                new_window['-COMBOBOXCATEGORY-'].update(values=table_values_category)
            elif new_event == '-ADDINCOME-':
                category_information = new_values['-COMBOCATEGORY-']
                convert_to_string = ' '.join(category_information)
                new_category_value = convert_to_string
                res = check_if_category_exist(new_category_value,table_values_category)
                #print(res)
                if res == 'does not exist':
                    sg.popup('Category does not exist, add the category to continue')
                    new_window.close()
                    window.un_hide()
                    break
                #print(f'new category info: {new_category_value}')
                detail_information = new_values['-INDETAIL-']
                amount_information = new_values['-INAMOUNT-']
                print(f'the category selected is: {new_category_value}, detail information is: {detail_information}, amount of this expense is: {amount_information}')
                register(detail_information,new_category_value,'Income',amount_information)
                new_window['-INDETAIL-'].update('')
                new_window['-INAMOUNT-'].update('')
                new_window['-COMBOCATEGORY-'].update(values=table_values_category)
        
        table_values_test = load_table()
        window['-MAINTABLE-'].update(values=table_values_test)


    window.close()

if __name__ == '__main__':
    main()